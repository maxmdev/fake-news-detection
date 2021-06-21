#%%

from __future__ import division, print_function

import os

from flask import Flask, request, render_template, redirect, url_for, send_from_directory, jsonify
from flask_wtf import Form
from wtforms import StringField, SubmitField, TextAreaField
from wtforms.validators import Required

import tokenizer as tokenizer
from nltk.tokenize.treebank import TreebankWordDetokenizer

seed_value = 0
os.environ['PYTHONHASHSEED'] = str(seed_value)
os.environ['CUDA_VISIBLE_DEVICES']='-1'
os.environ['TF_CUDNN_USE_AUTOTUNE'] ='0'
os.environ["KERAS_BACKEND"] = "plaidml.keras.backend"
os.environ['TFHUB_CACHE_DIR'] = './cache'

import random
random.seed(seed_value)

import numpy as np
from numpy.random import seed
seed(seed_value)

from pygments.lexers import go
from sklearn.feature_extraction.text import CountVectorizer

from tensorflow.python.framework import ops
from tensorflow.python.keras.backend import set_session
from sklearn.model_selection import train_test_split
import tensorflow as tf
import tensorflow._api.v2.compat.v1 as tf
from tensorflow._api.v2.compat.v1 import set_random_seed
set_random_seed(seed_value)

import tensorflow_hub as hub
import tensorflow_text

import pandas as pd

from tensorflow.keras.layers import LSTM, Dense, RepeatVector, Masking, TimeDistributed, Conv1D
from tensorflow.keras.layers import InputLayer
from tensorflow.keras import layers

import matplotlib.pyplot as plt

from uk_stemmer import UkStemmer
from stop_words import get_stop_words

from openpyxl import load_workbook

from nltk.tokenize import word_tokenize

import seaborn as sns

import itertools

import nltk
nltk.download('punkt')

from gensim.parsing.preprocessing import remove_stopwords
from gensim.parsing.preprocessing import STOPWORDS

import ssl
ssl._create_default_https_context = ssl._create_unverified_context

try:
    # Disable all GPUS
    tf.config.set_visible_devices([], 'GPU')
    visible_devices = tf.config.get_visible_devices()
    for device in visible_devices:
        assert device.device_type != 'GPU'
except:
    # Invalid device or cannot modify virtual devices once initialized.
    pass

tf.compat.v1.disable_eager_execution()
ops.reset_default_graph()

tf_config = ''
sess = tf.compat.v1.Session()

global graph
graph = tf.get_default_graph()
def reset_seeds():
   np.random.seed(seed_value)
   random.seed(seed_value)
   tf.random.set_random_seed(seed_value)
### Read data

print(tf.__version__)
reset_seeds()

# WebApp
app = Flask("__name__")
app.config['SECRET_KEY'] = 'hard to guess string'
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# Setting stop-words parameters
stop_words = get_stop_words('ukrainian', True)
my_stop_words = STOPWORDS.union(set(stop_words))

# Setting Stemmer
stemmer = UkStemmer()

# Set data classes
classes = ['Real', 'Fake']

# Loading trained model
set_session(sess)
model = tf.keras.models.load_model('.../saved_model/my_model')

# Check model summary
# model.summary()

# Load word embedding module
module_url = "https://tfhub.dev/google/universal-sentence-encoder-multilingual/3"
embed = hub.load(module_url)

# Get input embeddings function declare
def getEmbeddings(newsText):
    with tf.compat.v1.Session() as session:
        session.run([tf.compat.v1.global_variables_initializer(), tf.compat.v1.tables_initializer()])
        training_embeddings = session.run(embed([newsText]))
        return training_embeddings

# Model prediction function declare
def getPrediction(text):
    prediction_result = model.predict(getEmbeddings(text)[0:1])
    result = (prediction_result > 0.6).astype(int)
    return result

# Removing stop-words function declare
def stopWords(text):
    tokenizer = nltk.RegexpTokenizer(r"\w+")
    text_tokens = tokenizer.tokenize(text)
    tokens_without_sw = [word for word in text_tokens if not word in my_stop_words]
    filtered_text = TreebankWordDetokenizer().detokenize(tokens_without_sw)
    return filtered_text

# Stemming input function declare
def stemWords(text):
        text = [stemmer.stem_word(word) for word in text.split()]
        return " ".join(text)

# Preprocessing input text function declare
def textPreprocess(text):
        text = stopWords(text)
        text = stemWords(text)
        return text

# Get prediction label function declare
def getPredictionLabel(prediction):
        MaxPosition = np.argmax(prediction)
        prediction_label = classes[MaxPosition]
        return prediction_label

# Input data
# newsText = input('Enter newsline: ')

# Preprocessing text
# newsText = textPreprocess(newsText)

# print('Preprocessed text is: ' + newsText)

# Prediction call
# pred = getPrediction(newsText)

# Prediction label
# label = getPredictionLabel(pred)

# Global vars

# print(label)
class postForm(Form):
    title = StringField('Заголовок новини ', validators=[Required()])
    text = TextAreaField('Текст новини ', validators=[Required()])
    url = StringField('URL новини ', validators=[Required()])
    submit = SubmitField('Перевірити')
    upload = SubmitField('Додати до бази даних')

class operatorForm(Form):
    title = StringField('Заголовок новини: ', validators=[Required()])
    text = StringField('Текст новини: ', validators=[Required()])
    url = StringField('URL новини: ', validators=[Required()])
    label = StringField('Помітка новини: ', validators=[Required()])
    submitAccept = SubmitField('Submit')
    submitDecline = SubmitField('Submit')


@app.route('/images/<path:path>')
def send_img(path):
    return send_from_directory('./templates/images', path)


@app.route('/styles/<path:path>')
def send_csw(path):
    return send_from_directory('./templates/styles', path)


@app.route('/js/<path:path>')
def send_js(path):
    return send_from_directory('./templates/js', path)


@app.route("/", methods=['GET'])
def main():
    title = None
    text = None
    url = None
    form = postForm()
    return render_template('main.html', form=form, title=title, text=text, url=url)


@app.route("/prediction", methods=['GET'])
def go_main():
    return redirect('/')


@app.route("/prediction", methods=['POST'])
def send():
    title = None
    text = None
    url = None
    form = postForm()
    if form.validate_on_submit():
        form.title.data = ''
        form.text.data = ''
        form.url.data = ''

    source = form.text.data
    with graph.as_default():
        set_session(sess)
        form.text.data = textPreprocess(form.text.data)
        prediction = getPrediction(form.text.data)
        prediction = getPredictionLabel(prediction)
    return render_template('main2.html', form=form, title=form.title.data, text=form.text.data, url=form.url.data,
                           label=prediction, source=source)


@app.route("/approve", methods=['POST', 'GET'])
def approve():
    form = operatorForm()
    if form.validate_on_submit():
        form.title.data = ''
        form.text.data = ''
        form.url.data = ''
        form.submitAccept.data = ''
        form.submitDecline.data = ''
        form.label.data = ''

    if form.submitAccept.data:
        label = form.label.data
    if form.submitDecline.data:
        if form.label.data == 'Fake':
            label = 'Real'
        else:
            label = 'Fake'

    df = pd.DataFrame({'title': [form.title.data], 'text': [form.text.data], 'label': [label], 'url': [form.url.data]})
    writer = pd.ExcelWriter('verified_dataset.xlsx', engine='openpyxl')
    book = load_workbook('verified_dataset.xlsx')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    startrow = writer.sheets['ApprovedNews'].max_row

    for sheetname in writer.sheets:
     df.to_excel(writer, index=False, sheet_name='ApprovedNews', header=None, startrow=startrow)

    writer.save()
    writer.close()
    return go_main()

app.run(port=5000)
# print(newsPredict(newsText))